from openpyxl import load_workbook, Workbook
import os

file_names = os.listdir('data-input')
bugged_rows = []
target_questions = {'"question": "Diverse Desires - Intro"',
                    '"question": "Diverse Desires - Outro"',
                    '"question": "Diverse Beliefs - Intro"',
                    '"question": "Diverse Beliefs - Outro"',
                    '"question": "False Belief - Intro"',
                    '"question": "Hidden Emotions - Intro"',
                    '"question": "False Belief(Contents) - Mechanics Intro"'
                    '"question": "False Belief(Contents) - Outro"'
                    }
for i in file_names:
    if i != '.gitignore':
        wb = load_workbook(filename='data-input/' + i)
        original_worksheet = wb.active
        for row in original_worksheet.iter_rows():
            start_timestamp = row[4].value.split(": ")[1]
            end_timestamp = row[5].value.split(": ")[1]
            question_type = row[3].value.strip()
            print(question_type)
            if start_timestamp == end_timestamp and question_type in target_questions:
                bugged_rows.append(row)
        for row in bugged_rows:
            original_worksheet.delete_rows(row[0].row)
        sheet_with_bugged_times = wb.create_sheet("bugged_rows")
        for row in bugged_rows:
            row_copy = [i.value for i in row]
            sheet_with_bugged_times.append(row_copy)
        wb.save('data-output/' + i)
